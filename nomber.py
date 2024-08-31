import os
import openpyxl as xl
import pandas as pd
from tqdm import tqdm , trange
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import pandas as pd
from selenium.webdriver.chrome.service import Service


wb = xl.load_workbook('nimirofon.xlsx') 
sheet=wb[wb.sheetnames[0]]
listfon=[]
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,1).value!=None:
        nmrotl=sheet.cell(i,1).value
        listfon.append(nmrotl)
    else:
        continue

ser = Service("chromedriver.exe")
op = webdriver.ChromeOptions()
wrb = webdriver.Chrome(service=ser, options=op)

for iii in listfon:
    wrb.get('https://web.whatsapp.com/')
    time.sleep(20)
    try:
        while True:
            wrb.find_element(by=By.XPATH,value='/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/div')
            wrb.get('https://web.whatsapp.com/')
            time.sleep(20)
    except:
        try:
            wrb.find_element(by=By.XPATH,value='/html/body/div[1]/div/div/div[3]/div/header/div[1]/div/div/span')
        except:
            time.sleep(20)
        url="https://web.whatsapp.com/send?phone=%2B212610358508&text=J%27ai+une+question+concernant+le+logement+propos%C3%A9%60&type=phone_number&app_absent=0"
        wrb.get(url)
        time.sleep(5)
        wrb.find_element(by=By.XPATH,value='/html/body/div[1]/div/div/div[4]/div/footer/div[1]/div/span[2]/div/div[2]/div[2]/button').click()#send
        time.sleep(2)